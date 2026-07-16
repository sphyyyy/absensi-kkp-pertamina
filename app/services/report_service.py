import os
from datetime import date

from flask import current_app
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.models import Attendance, User
from app.utils.helpers import format_date, format_time


def _ensure_reports_dir():
    """Ensure the reports directory exists."""
    reports_dir = os.path.join(
        current_app.root_path, '..', 'reports'
    )
    os.makedirs(reports_dir, exist_ok=True)
    return os.path.abspath(reports_dir)


def generate_pdf_report(start_date, end_date, title=None):
    """Generate a PDF attendance report for a date range.

    Returns:
        str: Absolute path to the generated PDF file.
    """
    reports_dir = _ensure_reports_dir()
    filename = f'laporan_absensi_{start_date}_{end_date}.pdf'
    filepath = os.path.join(reports_dir, filename)

    if title is None:
        title = f'Laporan Absensi KKP — {format_date(start_date)} s/d {format_date(end_date)}'

    doc = SimpleDocTemplate(filepath, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=20*mm, bottomMargin=20*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Title'],
        fontSize=14, spaceAfter=12, alignment=1,
    )
    subtitle_style = ParagraphStyle(
        'Subtitle', parent=styles['Normal'],
        fontSize=10, spaceAfter=20, alignment=1,
        textColor=colors.grey,
    )

    elements = []

    # Title
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(
        'PT. Pertamina Patra Niaga Regional Sulawesi', subtitle_style
    ))
    elements.append(Spacer(1, 10))

    # Fetch data
    records = (
        Attendance.query
        .join(User)
        .filter(
            Attendance.attendance_date >= start_date,
            Attendance.attendance_date <= end_date,
            Attendance.is_valid == True  # noqa: E712
        )
        .order_by(Attendance.attendance_date, User.full_name)
        .all()
    )

    # Table header
    data = [['No', 'Nama', 'NIM', 'Tanggal', 'Masuk', 'Pulang', 'Status']]

    for i, rec in enumerate(records, 1):
        data.append([
            str(i),
            rec.user.full_name,
            rec.user.nim or '-',
            format_date(rec.attendance_date),
            format_time(rec.check_in_time),
            format_time(rec.check_out_time),
            rec.status.capitalize(),
        ])

    if len(data) == 1:
        data.append(['', '', '', 'Tidak ada data', '', '', ''])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003D7A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4F8')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    doc.build(elements)
    return filepath


def generate_excel_report(start_date, end_date):
    """Generate an Excel attendance report for a date range.

    Returns:
        str: Absolute path to the generated Excel file.
    """
    reports_dir = _ensure_reports_dir()
    filename = f'laporan_absensi_{start_date}_{end_date}.xlsx'
    filepath = os.path.join(reports_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Rekap Absensi'

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='003D7A', end_color='003D7A', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Title row
    ws.merge_cells('A1:G1')
    ws['A1'] = f'Laporan Absensi KKP — {format_date(start_date)} s/d {format_date(end_date)}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:G2')
    ws['A2'] = 'PT. Pertamina Patra Niaga Regional Sulawesi'
    ws['A2'].font = Font(size=10, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Header row
    headers = ['No', 'Nama', 'NIM', 'Tanggal', 'Jam Masuk', 'Jam Pulang', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    records = (
        Attendance.query
        .join(User)
        .filter(
            Attendance.attendance_date >= start_date,
            Attendance.attendance_date <= end_date,
            Attendance.is_valid == True  # noqa: E712
        )
        .order_by(Attendance.attendance_date, User.full_name)
        .all()
    )

    for i, rec in enumerate(records, 1):
        row = i + 4
        values = [
            i,
            rec.user.full_name,
            rec.user.nim or '-',
            format_date(rec.attendance_date),
            format_time(rec.check_in_time),
            format_time(rec.check_out_time),
            rec.status.capitalize(),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    # Auto-width columns
    col_widths = [5, 25, 15, 18, 12, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = width

    wb.save(filepath)
    return filepath
